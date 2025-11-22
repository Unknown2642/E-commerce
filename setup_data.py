import openpyxl
from openpyxl import Workbook
from werkzeug.security import generate_password_hash
import uuid
import os
import base64

def setup_sample_data():
    """Create sample data for the application"""
    
    # Create data directory if it doesn't exist
    if not os.path.exists('data'):
        os.makedirs('data')
    
    # Setup Users
    users_file = 'data/users.xlsx'
    if not os.path.exists(users_file):
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Email', 'Password', 'Role', 'Address', 'Phone', 'SecurityQuestion', 'SecurityAnswer'])
        
        # Add admin user
        admin_id = str(uuid.uuid4())
        admin_password = generate_password_hash('admin123')
        ws.append([admin_id, 'Admin User', 'admin@ecommerce.com', admin_password, 'admin', '123 Admin St, City', '555-0001', 'What is your favorite color?', 'blue'])
        
        # Add regular user
        user_id = str(uuid.uuid4())
        user_password = generate_password_hash('user123')
        ws.append([user_id, 'John Doe', 'user@ecommerce.com', user_password, 'user', '456 User Ave, City', '555-0002', 'What is your pet name?', 'buddy'])
        
        wb.save(users_file)
        print("Users created - Admin: admin@ecommerce.com / admin123, User: user@ecommerce.com / user123")
    
    # Setup Products
    products_file = 'data/products.xlsx'
    if not os.path.exists(products_file):
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Description', 'Price', 'Quantity', 'Image', 'Category'])
        
        # Create placeholder images
        import base64
        placeholder_img = 'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNkYPhfDwAChAHGbveP9gAAAABJRU5ErkJggg=='
        
        products = [
            ['Laptop', 'High-performance laptop for work and gaming', 45999.99, 15, 'images/laptop.png', 'Electronics'],
            ['Smartphone', 'Latest smartphone with advanced features', 25999.99, 20, 'images/phone.png', 'Electronics'],
            ['T-Shirt', 'Comfortable cotton t-shirt', 599.99, 100, 'images/tshirt.png', 'Clothing'],
            ['Jeans', 'Classic blue jeans', 1299.99, 75, 'images/jeans.png', 'Clothing'],
            ['Book', 'Bestselling novel', 299.99, 50, 'images/book.png', 'Books'],
            ['Coffee Mug', 'Ceramic coffee mug', 199.99, 80, 'images/mug.png', 'Home']
        ]
        
        # Create placeholder image files
        image_names = ['laptop.png', 'phone.png', 'tshirt.png', 'jeans.png', 'book.png', 'mug.png']
        for img_name in image_names:
            img_path = f'static/images/{img_name}'
            if not os.path.exists(img_path):
                with open(img_path, 'wb') as f:
                    f.write(base64.b64decode(placeholder_img))
        
        for product in products:
            product_id = str(uuid.uuid4())
            ws.append([product_id] + product)
        
        wb.save(products_file)
        print("Sample products created")
    
    # Setup empty Orders, Cart and Reviews files
    for file_name, headers in [
        ('data/orders.xlsx', ['ID', 'UserID', 'ProductID', 'Quantity', 'PaymentMethod', 'Status', 'Date']),
        ('data/cart.xlsx', ['ID', 'UserID', 'ProductID', 'Quantity']),
        ('data/reviews.xlsx', ['ID', 'UserID', 'ProductID', 'Rating', 'Review', 'Date'])
    ]:
        if not os.path.exists(file_name):
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            wb.save(file_name)
    
    print("Orders and Cart files initialized")
    print("\nSetup complete! You can now run the application.")
    print("Login credentials:")
    print("   Admin: admin@ecommerce.com / admin123")
    print("   User:  user@ecommerce.com / user123")

if __name__ == '__main__':
    setup_sample_data()