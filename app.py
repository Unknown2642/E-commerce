from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask_wtf.csrf import CSRFProtect, validate_csrf
from flask_wtf import FlaskForm
import openpyxl
from openpyxl import Workbook
import os
import uuid
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'
CORS(app)
csrf = CSRFProtect(app)

@app.after_request
def add_security_headers(response):
    if request.endpoint == 'checkout':
        return response
    
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline'; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; font-src 'self' https://cdn.jsdelivr.net"
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    return response

limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["20 per minute"]
)

UPLOAD_FOLDER = 'static/images'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def init_excel_files():
    if not os.path.exists('data'):
        os.makedirs('data')
    
    files_exist = all(os.path.exists(f) for f in [
        'data/users.xlsx', 'data/products.xlsx', 'data/orders.xlsx', 'data/cart.xlsx', 'data/reviews.xlsx'
    ])
    
    if not files_exist:
        from setup_data import setup_sample_data
        setup_sample_data()

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data = []
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            data.append(dict(zip(headers, row)))
    return data

def write_excel(file_path, data):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.append(data)
    wb.save(file_path)

def update_excel(file_path, row_id, updated_data):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == row_id:
            for col_num, header in enumerate(headers, start=1):
                if header in updated_data:
                    ws.cell(row=row_num, column=col_num, value=updated_data[header])
            break
    wb.save(file_path)

def get_cart_count():
    if 'user_id' not in session:
        return 0
    
    cart_items = read_excel('data/cart.xlsx')
    user_cart = [item for item in cart_items if item['UserID'] == session['user_id']]
    return len(user_cart)

app.jinja_env.globals.update(get_cart_count=get_cart_count)

@app.route('/')
def home():
    products = read_excel('data/products.xlsx')
    return render_template('home.html', products=products[:6])

@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def register():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        phone = request.form['phone']
        address = request.form['address']
        security_question = request.form['security_question']
        security_answer = request.form['security_answer']
        
        users = read_excel('data/users.xlsx')
        if any(user['Email'] == email for user in users):
            flash('Email already exists')
            return render_template('register.html')
        
        user_id = str(uuid.uuid4())
        hashed_password = generate_password_hash(password)
        write_excel('data/users.xlsx', [user_id, name, email, hashed_password, 'user', address, phone, security_question, security_answer])
        
        flash('Registration successful')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
@limiter.exempt
def login():
    if request.method == 'POST':
        login_input = request.form['email']
        password = request.form['password']
        
        users = read_excel('data/users.xlsx')
        user = next((u for u in users if u['Email'] == login_input or u['Name'] == login_input), None)
        
        if user and check_password_hash(user['Password'], password):
            session['user_id'] = user['ID']
            session['user_name'] = user['Name']
            session['user_role'] = user['Role']
            
            if user['Role'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('dashboard'))
        
        flash('Invalid credentials')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('home'))

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email_or_name = request.form['email_or_name']
        
        users = read_excel('data/users.xlsx')
        user = next((u for u in users if u['Email'] == email_or_name or u['Name'] == email_or_name), None)
        
        if user and user.get('SecurityQuestion'):
            session['reset_user_id'] = user['ID']
            return redirect(url_for('security_question'))
        else:
            flash('User not found or no security question set')
    
    return render_template('forgot_password.html')

@app.route('/security_question', methods=['GET', 'POST'])
def security_question():
    if 'reset_user_id' not in session:
        return redirect(url_for('forgot_password'))
    
    users = read_excel('data/users.xlsx')
    user = next((u for u in users if u['ID'] == session['reset_user_id']), None)
    
    if not user:
        session.pop('reset_user_id', None)
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        answer = request.form['answer'].strip().lower()
        correct_answer = user.get('SecurityAnswer', '').strip().lower()
        
        if answer == correct_answer:
            return redirect(url_for('reset_password'))
        else:
            flash('Incorrect answer. Please try again.')
    
    return render_template('security_question.html', question=user.get('SecurityQuestion'))

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if 'reset_user_id' not in session:
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        new_password = request.form['password']
        confirm_password = request.form['confirm_password']
        
        if new_password != confirm_password:
            flash('Passwords do not match')
            return render_template('reset_password.html')
        
        hashed_password = generate_password_hash(new_password)
        update_excel('data/users.xlsx', session['reset_user_id'], {'Password': hashed_password})
        
        session.pop('reset_user_id', None)
        flash('Password reset successfully. Please login with your new password.')
        return redirect(url_for('login'))
    
    return render_template('reset_password.html')

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session['user_role'] == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    products = read_excel('data/products.xlsx')
    return render_template('dashboard.html', products=products)

@app.route('/profile')
def profile():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    users = read_excel('data/users.xlsx')
    user = next((u for u in users if u['ID'] == session['user_id']), None)
    
    # Get user's orders
    orders = read_excel('data/orders.xlsx')
    products = read_excel('data/products.xlsx')
    
    user_orders = []
    for order in orders:
        if order.get('UserID') == session['user_id']:
            product = next((p for p in products if p['ID'] == order['ProductID']), {})
            user_orders.append({
                'order': order,
                'product': product
            })
    
    # Sort orders by date (newest first)
    user_orders.sort(key=lambda x: x['order'].get('Date', ''), reverse=True)
    
    return render_template('profile.html', user=user, orders=user_orders)

@app.route('/update_profile', methods=['POST'])
def update_profile():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    phone = request.form['phone']
    address = request.form['address']
    
    update_excel('data/users.xlsx', session['user_id'], {'Phone': phone, 'Address': address})
    flash('Profile updated successfully')
    return redirect(url_for('profile'))

@app.route('/product/<product_id>')
def product_detail(product_id):
    products = read_excel('data/products.xlsx')
    product = next((p for p in products if p['ID'] == product_id), None)
    if not product:
        return "Product not found", 404
    
    reviews = read_excel('data/reviews.xlsx')
    users = read_excel('data/users.xlsx')
    
    product_reviews = []
    total_rating = 0
    
    for review in reviews:
        if review['ProductID'] == product_id:
            user = next((u for u in users if u['ID'] == review['UserID']), {})
            product_reviews.append({
                'review': review,
                'user_name': user.get('Name', 'Anonymous')
            })
            total_rating += int(review['Rating'])
    
    avg_rating = round(total_rating / len(product_reviews), 1) if product_reviews else 0
    
    return render_template('product_detail.html', 
                         product=product, 
                         reviews=product_reviews,
                         avg_rating=avg_rating,
                         review_count=len(product_reviews))

@app.route('/add_to_cart', methods=['POST'])
def add_to_cart():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    product_id = request.form['product_id']
    quantity = int(request.form['quantity'])
    
    products = read_excel('data/products.xlsx')
    product = next((p for p in products if p['ID'] == product_id), None)
    
    if not product:
        flash('Product not found')
        return redirect(url_for('dashboard'))
    
    available_stock = int(product['Quantity'])
    
    if available_stock <= 0:
        flash(f'{product["Name"]} is out of stock')
        return redirect(url_for('product_detail', product_id=product_id))
    
    if quantity > available_stock:
        flash(f'Only {available_stock} items available for {product["Name"]}')
        return redirect(url_for('product_detail', product_id=product_id))
    
    cart_id = str(uuid.uuid4())
    write_excel('data/cart.xlsx', [cart_id, session['user_id'], product_id, quantity])
    
    flash('Product added to cart')
    return redirect(url_for('dashboard'))

@app.route('/cart')
def cart():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    cart_items = read_excel('data/cart.xlsx')
    products = read_excel('data/products.xlsx')
    
    user_cart = [item for item in cart_items if item['UserID'] == session['user_id']]
    
    cart_with_products = []
    total_amount = 0.0
    
    for item in user_cart:
        product = next((p for p in products if p['ID'] == item['ProductID']), None)
        if product:
            try:
                item_total = float(product['Price']) * int(item['Quantity'])
                total_amount += item_total
                cart_with_products.append({
                    'cart_id': item['ID'],
                    'product': product,
                    'quantity': int(item['Quantity']),
                    'total': item_total
                })
            except (ValueError, TypeError):
                continue
    
    # Pass redirect parameter to template
    redirect_param = request.args.get('redirect', '')
    
    return render_template('cart.html', 
                         cart_items=cart_with_products, 
                         total_amount=total_amount,
                         redirect_param=redirect_param)

@app.route('/checkout', methods=['GET', 'POST'])
def checkout():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get cart items for display
    cart_items = read_excel('data/cart.xlsx')
    products = read_excel('data/products.xlsx')
    
    user_cart = [item for item in cart_items if item['UserID'] == session['user_id']]
    
    if not user_cart:
        flash('Your cart is empty')
        return redirect(url_for('cart'))
    
    cart_with_products = []
    subtotal = 0.0
    
    if request.method == 'POST' and request.form.get('subtotal'):
        try:
            subtotal = float(request.form.get('subtotal', 0))
        except:
            subtotal = 0.0
    
    for i, item in enumerate(user_cart, 1):
        product = next((p for p in products if p['ID'] == item['ProductID']), None)
        if product:
            try:
                if request.method == 'POST':
                    form_price = request.form.get(f'item_price_{i}')
                    form_quantity = request.form.get(f'item_quantity_{i}')
                    form_total = request.form.get(f'item_total_{i}')
                    
                    if form_price and form_quantity and form_total:
                        item_price = float(form_price)
                        item_quantity = int(form_quantity)
                        item_total = float(form_total)
                    else:
                        item_price = float(product['Price'])
                        item_quantity = int(item['Quantity'])
                        item_total = item_price * item_quantity
                else:
                    item_price = float(product['Price'])
                    item_quantity = int(item['Quantity'])
                    item_total = item_price * item_quantity
                    subtotal += item_total
                
                cart_with_products.append({
                    'product': product,
                    'quantity': item_quantity,
                    'total': item_total
                })
            except (ValueError, TypeError):
                continue
    
    discount_amount = 0.0
    coupon_code = request.args.get('coupon', '') or request.form.get('coupon_code', '')
    coupon_code = coupon_code.strip().upper()
    coupon_valid = False
    
    if request.method == 'POST' and request.form.get('discount_amount'):
        try:
            discount_amount = float(request.form.get('discount_amount', 0))
            coupon_valid = True
        except:
            discount_amount = 0.0
    
    elif coupon_code:
        if coupon_code == 'SAVE10':
            discount_amount = subtotal * 0.10
            coupon_valid = True
        elif coupon_code == 'SAVE20':
            discount_amount = subtotal * 0.20
            coupon_valid = True
        elif coupon_code == 'WELCOME':
            discount_amount = min(50.0, subtotal * 0.05)
            coupon_valid = True
        elif coupon_code == 'SAVE100':
            discount_amount = subtotal * 1.0
            coupon_valid = True
        else:
            flash('Invalid coupon code. Please try again.', 'error')
    
    if request.method == 'POST':
        payment_method = request.form['payment_method']
        
        # Validate payment details based on method
        if payment_method == 'card':
            card_number = request.form.get('card_number', '').strip()
            expiry = request.form.get('expiry', '').strip()
            cvv = request.form.get('cvv', '').strip()
            
            if not card_number or not expiry or not cvv:
                flash('Please fill all card details')
                return render_template('checkout.html', 
                                     cart_items=cart_with_products, 
                                     subtotal=subtotal,
                                     discount_amount=discount_amount,
                                     coupon_code=coupon_code if coupon_valid else '',
                                     total_amount=subtotal - discount_amount)
        
        elif payment_method == 'upi':
            upi_id = request.form.get('upi_id', '').strip()
            upi_pin = request.form.get('upi_pin', '').strip()
            
            if not upi_id or not upi_pin:
                flash('Please fill all UPI details')
                return render_template('checkout.html', 
                                     cart_items=cart_with_products, 
                                     subtotal=subtotal,
                                     discount_amount=discount_amount,
                                     coupon_code=coupon_code if coupon_valid else '',
                                     total_amount=subtotal - discount_amount)
        
        # Final stock validation before order placement
        for item in user_cart:
            product = next((p for p in products if p['ID'] == item['ProductID']), None)
            if product:
                available_stock = int(product['Quantity'])
                ordered_quantity = int(item['Quantity'])
                
                if available_stock <= 0:
                    flash(f'{product["Name"]} is out of stock. Please remove from cart.')
                    return render_template('checkout.html', 
                                         cart_items=cart_with_products, 
                                         subtotal=subtotal,
                                         discount_amount=discount_amount,
                                         coupon_code=coupon_code if coupon_valid else '',
                                         total_amount=subtotal - discount_amount)
                
                if ordered_quantity > available_stock:
                    flash(f'Only {available_stock} items available for {product["Name"]}. Please update quantity.')
                    return render_template('checkout.html', 
                                         cart_items=cart_with_products, 
                                         subtotal=subtotal,
                                         discount_amount=discount_amount,
                                         coupon_code=coupon_code if coupon_valid else '',
                                         total_amount=subtotal - discount_amount)
        
        final_total = subtotal - discount_amount
        
        final_total = subtotal - discount_amount
        
        for item in user_cart:
            order_id = str(uuid.uuid4())
            write_excel('data/orders.xlsx', [
                order_id, session['user_id'], item['ProductID'], 
                item['Quantity'], payment_method, 'pending', datetime.now().isoformat()
            ])
            
            # Don't deduct stock here - wait for admin approval
        
        # Clear cart
        wb = openpyxl.load_workbook('data/cart.xlsx')
        ws = wb.active
        rows_to_delete = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[1].value == session['user_id']:
                rows_to_delete.append(row_num)
        
        for row_num in reversed(rows_to_delete):
            ws.delete_rows(row_num)
        wb.save('data/cart.xlsx')
        
        flash('Order placed successfully')
        
        # 
        redirect_url = request.form.get('redirect') or request.args.get('redirect', url_for('dashboard'))
        print(f"Debug: Redirecting to: {redirect_url}")  # Debug output
        return redirect(redirect_url)
    
    # 
    redirect_param = request.args.get('redirect', '')
    
    return render_template('checkout.html', 
                         cart_items=cart_with_products, 
                         subtotal=subtotal,
                         discount_amount=discount_amount,
                         coupon_code=coupon_code if coupon_valid else '',
                         total_amount=subtotal - discount_amount,
                         redirect_param=redirect_param)

@app.route('/update_cart', methods=['POST'])
def update_cart():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    cart_id = request.form['cart_id']
    new_quantity = int(request.form['quantity'])
    
    if new_quantity <= 0:
        flash('Invalid quantity')
        return redirect(url_for('cart'))
    
    # 
    
    update_excel('data/cart.xlsx', cart_id, {'Quantity': new_quantity})
    flash('Cart updated successfully')
    return redirect(url_for('cart'))

@app.route('/remove_from_cart', methods=['POST'])
def remove_from_cart():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    cart_id = request.form['cart_id']
    
    # Remove item from cart
    wb = openpyxl.load_workbook('data/cart.xlsx')
    ws = wb.active
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == cart_id:
            ws.delete_rows(row_num)
            break
    wb.save('data/cart.xlsx')
    
    flash('Item removed from cart')
    return redirect(url_for('cart'))

# Admin routes
@app.route('/admin')
def admin_dashboard():
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    users = read_excel('data/users.xlsx')
    products = read_excel('data/products.xlsx')
    orders = read_excel('data/orders.xlsx')
    
    return render_template('admin_dashboard.html', 
                         users=len(users), 
                         products=len(products), 
                         orders=len(orders))

@app.route('/admin/users')
def admin_users():
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    users = read_excel('data/users.xlsx')
    total = len(users)
    start = (page - 1) * per_page
    end = start + per_page
    users_page = users[start:end]
    
    total_pages = (total + per_page - 1) // per_page
    
    return render_template('admin_users.html', 
                         users=users_page, 
                         page=page, 
                         total_pages=total_pages,
                         total=total)

@app.route('/admin/products')
def admin_products():
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    products = read_excel('data/products.xlsx')
    total = len(products)
    start = (page - 1) * per_page
    end = start + per_page
    products_page = products[start:end]
    
    total_pages = (total + per_page - 1) // per_page
    
    return render_template('admin_products.html', 
                         products=products_page, 
                         page=page, 
                         total_pages=total_pages,
                         total=total)

@app.route('/admin/add_product', methods=['GET', 'POST'])
def add_product():
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        name = request.form['name']
        description = request.form['description']
        price = float(request.form['price'])
        quantity = int(request.form['quantity'])
        category = request.form['category']
        
        image_path = ''
        if 'image' in request.files:
            file = request.files['image']
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filename = f"{uuid.uuid4()}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                image_path = f"images/{filename}"
        
        product_id = str(uuid.uuid4())
        write_excel('data/products.xlsx', [product_id, name, description, price, quantity, image_path, category])
        
        flash('Product added successfully')
        return redirect(url_for('admin_products'))
    
    return render_template('add_product.html')

@app.route('/admin/orders')
def admin_orders():
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    page = request.args.get('page', 1, type=int)
    query = request.args.get('q', '')
    status_filter = request.args.get('status', '')
    date_filter = request.args.get('date', '')
    payment_filter = request.args.get('payment', '')
    per_page = 10
    
    orders = read_excel('data/orders.xlsx')
    users = read_excel('data/users.xlsx')
    products = read_excel('data/products.xlsx')
    
    # Apply all filters
    filtered_orders = []
    for order in orders:
        user = next((u for u in users if u['ID'] == order['UserID']), {})
        product = next((p for p in products if p['ID'] == order['ProductID']), {})
        
        # Handle deleted users
        if not user:
            user = {
                'Name': 'Deleted User',
                'Email': 'deleted@user.com',
                'ID': order['UserID']
            }
        
        # Text search filter
        if query:
            search_match = (query.lower() in order.get('ID', '').lower() or
                          query.lower() in user.get('Name', '').lower() or
                          query.lower() in user.get('Email', '').lower() or
                          query.lower() in product.get('Name', '').lower() or
                          query.lower() in order.get('Status', '').lower() or
                          query.lower() in order.get('PaymentMethod', '').lower() or
                          query.lower() in order.get('Date', '').lower())
            if not search_match:
                continue
        
        # Status filter
        if status_filter and order['Status'] != status_filter:
            continue
        
        # Date filter
        if date_filter and order.get('Date', '')[:10] != date_filter:
            continue
        
        # Payment method filter
        if payment_filter and order['PaymentMethod'] != payment_filter:
            continue
        
        filtered_orders.append(order)
    
    # Sort by date (newest first)
    filtered_orders.sort(key=lambda x: x.get('Date', ''), reverse=True)
    
    # Pagination
    total = len(filtered_orders)
    start = (page - 1) * per_page
    end = start + per_page
    orders_page = filtered_orders[start:end]
    
    orders_with_details = []
    for order in orders_page:
        user = next((u for u in users if u['ID'] == order['UserID']), {})
        product = next((p for p in products if p['ID'] == order['ProductID']), {})
        orders_with_details.append({
            'order': order,
            'user': user,
            'product': product
        })
    
    total_pages = (total + per_page - 1) // per_page
    
    return render_template('admin_orders.html', 
                         orders=orders_with_details,
                         page=page,
                         total_pages=total_pages,
                         total=total,
                         query=query,
                         status_filter=status_filter,
                         date_filter=date_filter,
                         payment_filter=payment_filter)

@app.route('/admin/add_user', methods=['GET', 'POST'])
def admin_add_user():
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        allowed_fields = {'name', 'email', 'password', 'phone', 'address', 'security_question', 'security_answer', 'role', 'csrf_token'}
        received_fields = set(request.form.keys())
        extra_fields = received_fields - allowed_fields
        
        if extra_fields:
            flash(f'Invalid parameters detected: {", ".join(extra_fields)}')
            return render_template('admin_add_user.html')
        
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        phone = request.form['phone']
        address = request.form['address']
        
        users = read_excel('data/users.xlsx')
        if any(user['Email'] == email for user in users):
            flash('Email already exists')
            return render_template('admin_add_user.html')
        
        user_id = str(uuid.uuid4())
        hashed_password = generate_password_hash(password)
        security_question = request.form['security_question']
        security_answer = request.form['security_answer']
        
        write_excel('data/users.xlsx', [user_id, name, email, hashed_password, 'user', address, phone, security_question, security_answer])
        
        flash('User created successfully')
        return redirect(url_for('admin_users'))
    
    return render_template('admin_add_user.html')

@app.route('/search')
def search():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    query = request.args.get('q', '').lower()
    products = read_excel('data/products.xlsx')
    
    if query:
        filtered_products = [p for p in products if 
                           query in p['Name'].lower() or 
                           query in p['Description'].lower() or 
                           query in p['Category'].lower()]
    else:
        filtered_products = products
    
    return render_template('search_results.html', products=filtered_products, query=query)

@app.route('/admin/search_users')
def admin_search_users():
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    query = request.args.get('q', '').lower()
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    users = read_excel('data/users.xlsx')
    
    if query:
        filtered_users = [u for u in users if 
                         query in u['Name'].lower() or 
                         query in u['Email'].lower() or 
                         query in u['Role'].lower()]
    else:
        filtered_users = users
    
    total = len(filtered_users)
    start = (page - 1) * per_page
    end = start + per_page
    users_page = filtered_users[start:end]
    
    total_pages = (total + per_page - 1) // per_page
    
    return render_template('admin_users.html', 
                         users=users_page, 
                         query=query, 
                         page=page, 
                         total_pages=total_pages,
                         total=total)

@app.route('/admin/search_products')
def admin_search_products():
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    query = request.args.get('q', '').lower()
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    products = read_excel('data/products.xlsx')
    
    if query:
        filtered_products = [p for p in products if 
                           query in p['Name'].lower() or 
                           query in p['Description'].lower() or 
                           query in p['Category'].lower()]
    else:
        filtered_products = products
    
    total = len(filtered_products)
    start = (page - 1) * per_page
    end = start + per_page
    products_page = filtered_products[start:end]
    
    total_pages = (total + per_page - 1) // per_page
    
    return render_template('admin_products.html', 
                         products=products_page, 
                         query=query, 
                         page=page, 
                         total_pages=total_pages,
                         total=total)

@app.route('/admin/user/<user_id>')
def admin_user_detail(user_id):
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    users = read_excel('data/users.xlsx')
    user = next((u for u in users if u['ID'] == user_id), None)
    
    if not user:
        flash('User not found')
        return redirect(url_for('admin_users'))
    
    # Get user's orders
    orders = read_excel('data/orders.xlsx')
    products = read_excel('data/products.xlsx')
    
    user_orders = []
    for order in orders:
        if order['UserID'] == user_id:
            product = next((p for p in products if p['ID'] == order['ProductID']), {})
            user_orders.append({
                'order': order,
                'product': product
            })
    
    return render_template('admin_user_detail.html', user=user, orders=user_orders)

@app.route('/admin/edit_user/<user_id>', methods=['GET', 'POST'])
def admin_edit_user(user_id):
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    users = read_excel('data/users.xlsx')
    user = next((u for u in users if u['ID'] == user_id), None)
    
    if not user:
        flash('User not found')
        return redirect(url_for('admin_users'))
    
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        role = request.form['role']
        phone = request.form['phone']
        address = request.form['address']
        
        # Check if email already exists for other users
        if any(u['Email'] == email and u['ID'] != user_id for u in users):
            flash('Email already exists for another user')
            return render_template('admin_edit_user.html', user=user)
        
        # Update user data
        update_data = {
            'Name': name,
            'Email': email,
            'Role': role,
            'Phone': phone,
            'Address': address
        }
        
        update_excel('data/users.xlsx', user_id, update_data)
        flash('User updated successfully')
        return redirect(url_for('admin_user_detail', user_id=user_id))
    
    return render_template('admin_edit_user.html', user=user)

@app.route('/admin/view_password/<user_id>')
def admin_view_password(user_id):
    if 'user_id' not in session or session['user_role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    
    users = read_excel('data/users.xlsx')
    user = next((u for u in users if u['ID'] == user_id), None)
    
    if not user:
        return jsonify({'success': False, 'message': 'User not found'})
    
    # 
    mock_passwords = {
        'admin@ecommerce.com': 'admin123',
        'user@ecommerce.com': 'user123'
    }
    
    password = mock_passwords.get(user['Email'], 'password123')
    return jsonify({'success': True, 'password': password})

@app.route('/delete_user/<user_id>', methods=['POST'])
@csrf.exempt
def delete_user(user_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'})
    
  
    if user_id == session['user_id']:
        return jsonify({'success': False, 'message': 'Cannot delete your own account'})
    
    try:
        # Delete user from users.xlsx
        wb = openpyxl.load_workbook('data/users.xlsx')
        ws = wb.active
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == user_id:
                ws.delete_rows(row_num)
                break
        wb.save('data/users.xlsx')
        
        # Delete user's cart
        wb = openpyxl.load_workbook('data/cart.xlsx')
        ws = wb.active
        rows_to_delete = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[1].value == user_id:
                rows_to_delete.append(row_num)
        for row_num in reversed(rows_to_delete):
            ws.delete_rows(row_num)
        wb.save('data/cart.xlsx')
        
        
        if user_id == session['user_id']:
            session.clear()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/edit_product/<product_id>', methods=['GET', 'POST'])
def admin_edit_product(product_id):
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    products = read_excel('data/products.xlsx')
    product = next((p for p in products if p['ID'] == product_id), None)
    
    if not product:
        flash('Product not found')
        return redirect(url_for('admin_products'))
    
    if request.method == 'POST':
        name = request.form['name']
        description = request.form['description']
        price = float(request.form['price'])
        quantity = int(request.form['quantity'])
        category = request.form['category']
        
        image_path = product['Image']
        if 'image' in request.files:
            file = request.files['image']
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filename = f"{uuid.uuid4()}_{filename}"
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                image_path = f"images/{filename}"
        
        update_data = {
            'Name': name,
            'Description': description,
            'Price': price,
            'Quantity': quantity,
            'Image': image_path,
            'Category': category
        }
        
        update_excel('data/products.xlsx', product_id, update_data)
        flash('Product updated successfully')
        return redirect(url_for('admin_products'))
    
    return render_template('admin_edit_product.html', product=product)

@app.route('/admin/delete_product/<product_id>', methods=['POST'])
def admin_delete_product(product_id):
    if 'user_id' not in session or session['user_role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        # Delete product from products.xlsx
        wb = openpyxl.load_workbook('data/products.xlsx')
        ws = wb.active
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == product_id:
                ws.delete_rows(row_num)
                break
        wb.save('data/products.xlsx')
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/update_order_status', methods=['POST'])
@csrf.exempt
def admin_update_order_status():
    if 'user_id' not in session or session['user_role'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        data = request.get_json()
        order_id = data['order_id']
        new_status = data['status']
        
        if new_status not in ['pending', 'completed', 'cancelled']:
            return jsonify({'success': False, 'message': 'Invalid status'})
        
        # Get order details for validation
        orders = read_excel('data/orders.xlsx')
        order = next((o for o in orders if o['ID'] == order_id), None)
        
        if not order:
            return jsonify({'success': False, 'message': 'Order not found'})
        
        # If approving order, validate stock availability - HARD STOP
        if new_status == 'completed':
            products = read_excel('data/products.xlsx')
            product = next((p for p in products if p['ID'] == order['ProductID']), None)
            
            if not product:
                return jsonify({
                    'success': False, 
                    'message': 'Product not found. Cannot approve order.'
                })
            
            current_stock = int(product['Quantity'])
            ordered_quantity = int(order['Quantity'])
            
            # HARD STOP: Cannot approve if insufficient stock
            if current_stock < ordered_quantity:
                return jsonify({
                    'success': False, 
                    'message': f'APPROVAL BLOCKED: Insufficient stock for {product["Name"]}. Available: {current_stock}, Required: {ordered_quantity}. Cannot approve this order.'
                })
            
            # HARD STOP: Cannot approve if negative stock
            if current_stock <= 0:
                return jsonify({
                    'success': False, 
                    'message': f'APPROVAL BLOCKED: {product["Name"]} is out of stock (Stock: {current_stock}). Cannot approve this order.'
                })
            
            # Deduct stock only when approving
            new_quantity = current_stock - ordered_quantity
            update_excel('data/products.xlsx', order['ProductID'], {'Quantity': new_quantity})
        
        
        elif new_status == 'cancelled':
            if order:
                # 
                if order['PaymentMethod'] in ['card', 'upi']:
                    # Mock refund processing
                    products = read_excel('data/products.xlsx')
                    product = next((p for p in products if p['ID'] == order['ProductID']), None)
                    refund_amount = float(product['Price']) * order['Quantity'] if product else 0
                    # In real implementation: process_refund(order['UserID'], refund_amount, order['PaymentMethod'])
                    pass
        
        update_excel('data/orders.xlsx', order_id, {'Status': new_status})
        
        message = 'Order approved successfully' if new_status == 'completed' else 'Order cancelled and refund processed'
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/add_review/<product_id>', methods=['POST'])
def add_review(product_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    rating = int(request.form['rating'])
    review_text = request.form['review']
    
    # Check if user already reviewed this product
    reviews = read_excel('data/reviews.xlsx')
    existing_review = next((r for r in reviews if r['UserID'] == session['user_id'] and r['ProductID'] == product_id), None)
    
    if existing_review:
        # Update existing review
        update_excel('data/reviews.xlsx', existing_review['ID'], {
            'Rating': rating,
            'Review': review_text,
            'Date': datetime.now().isoformat()
        })
        flash('Review updated successfully')
    else:
        # Add new review
        review_id = str(uuid.uuid4())
        write_excel('data/reviews.xlsx', [
            review_id, session['user_id'], product_id, rating, review_text, datetime.now().isoformat()
        ])
        flash('Review added successfully')
    
    return redirect(url_for('product_detail', product_id=product_id))

@app.route('/admin/search_orders')
def admin_search_orders():
    if 'user_id' not in session or session['user_role'] != 'admin':
        return redirect(url_for('login'))
    
    query = request.args.get('q', '').lower()
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    orders = read_excel('data/orders.xlsx')
    users = read_excel('data/users.xlsx')
    products = read_excel('data/products.xlsx')
    
    # Search in orders, users, and products
    filtered_orders = []
    for order in orders:
        user = next((u for u in users if u['ID'] == order['UserID']), {})
        product = next((p for p in products if p['ID'] == order['ProductID']), {})
        
        if (query in order.get('ID', '').lower() or
            query in user.get('Name', '').lower() or
            query in user.get('Email', '').lower() or
            query in product.get('Name', '').lower() or
            query in order.get('Status', '').lower() or
            query in order.get('PaymentMethod', '').lower() or
            query in order.get('Date', '').lower()):
            filtered_orders.append(order)
    
    # Sort by date (newest first)
    filtered_orders.sort(key=lambda x: x.get('Date', ''), reverse=True)
    
    # Pagination
    total = len(filtered_orders)
    start = (page - 1) * per_page
    end = start + per_page
    orders_page = filtered_orders[start:end]
    
    orders_with_details = []
    for order in orders_page:
        user = next((u for u in users if u['ID'] == order['UserID']), {})
        product = next((p for p in products if p['ID'] == order['ProductID']), {})
        orders_with_details.append({
            'order': order,
            'user': user,
            'product': product
        })
    
    total_pages = (total + per_page - 1) // per_page
    
    return render_template('admin_orders.html', 
                         orders=orders_with_details,
                         query=query,
                         page=page,
                         total_pages=total_pages,
                         total=total)

@app.route('/api/validate_coupon', methods=['POST'])
@csrf.exempt
def api_validate_coupon():
    data = request.get_json()
    coupon_code = data.get('coupon_code', '').strip().upper()
    subtotal = float(data.get('subtotal', 0))
    
    discount_amount = 0.0
    valid = False
    message = ''
    
    if coupon_code == 'SAVE10':
        discount_amount = subtotal * 0.10
        valid = True
        message = '10% discount applied!'
    elif coupon_code == 'SAVE20':
        discount_amount = subtotal * 0.20
        valid = True
        message = '20% discount applied!'
    elif coupon_code == 'WELCOME':
        discount_amount = min(50.0, subtotal * 0.05)
        valid = True
        message = 'Welcome discount applied!'
    elif coupon_code == 'SAVE100':
        discount_amount = subtotal * 1.0
        valid = True
        message = '100% discount applied - FREE ORDER!'
    else:
        message = 'Invalid coupon code'
    
    return jsonify({
        'valid': valid,
        'discount_amount': discount_amount,
        'message': message,
        'final_total': subtotal - discount_amount
    })

@app.route('/api/process_checkout', methods=['POST'])
@csrf.exempt
def api_process_checkout():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Not logged in'})
    
    data = request.get_json()
    payment_method = data.get('payment_method')
    discount_amount = float(data.get('discount_amount', 0))
    
    # Get cart items
    cart_items = read_excel('data/cart.xlsx')
    products = read_excel('data/products.xlsx')
    user_cart = [item for item in cart_items if item['UserID'] == session['user_id']]
    
    if not user_cart:
        return jsonify({'success': False, 'message': 'Cart is empty'})
    
    # Validate payment details
    if payment_method == 'card':
        if not data.get('card_number') or not data.get('expiry') or not data.get('cvv'):
            return jsonify({'success': False, 'message': 'Please fill all card details'})
    elif payment_method == 'upi':
        if not data.get('upi_id') or not data.get('upi_pin'):
            return jsonify({'success': False, 'message': 'Please fill all UPI details'})
    
    # Stock validation
    for item in user_cart:
        product = next((p for p in products if p['ID'] == item['ProductID']), None)
        if product:
            if int(product['Quantity']) < int(item['Quantity']):
                return jsonify({
                    'success': False, 
                    'message': f'Insufficient stock for {product["Name"]}'
                })
    
    # Create orders
    for item in user_cart:
        order_id = str(uuid.uuid4())
        write_excel('data/orders.xlsx', [
            order_id, session['user_id'], item['ProductID'], 
            item['Quantity'], payment_method, 'pending', datetime.now().isoformat()
        ])
    
    # Clear cart
    wb = openpyxl.load_workbook('data/cart.xlsx')
    ws = wb.active
    rows_to_delete = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[1].value == session['user_id']:
            rows_to_delete.append(row_num)
    
    for row_num in reversed(rows_to_delete):
        ws.delete_rows(row_num)
    wb.save('data/cart.xlsx')
    
    return jsonify({
        'success': True, 
        'message': 'Order placed successfully',
        'redirect_url': data.get('redirect_url', url_for('dashboard'))
    })

@app.route('/api/validate_upi', methods=['POST'])
@csrf.exempt
def api_validate_upi():
    data = request.get_json()
    upi_id = data.get('upi_id', '').strip()
    
    # UPI ID validation patterns
    valid_patterns = [
        r'^[a-zA-Z0-9._-]+@[a-zA-Z0-9]+$',  # Basic UPI format
        r'^[0-9]{10}@[a-zA-Z0-9]+$',       # Mobile number based
        r'^[a-zA-Z0-9._-]+@(paytm|phonepe|gpay|bhim|ybl|okaxis|okhdfcbank|okicici|oksbi|axl)$'  # Popular providers
    ]
    
    is_valid = False
    provider = 'Unknown'
    
    if upi_id:
        import re
        for pattern in valid_patterns:
            if re.match(pattern, upi_id, re.IGNORECASE):
                is_valid = True
                break
        
        # Extract provider
        if '@' in upi_id:
            provider_code = upi_id.split('@')[1].lower()
            providers = {
                'paytm': 'Paytm',
                'phonepe': 'PhonePe', 
                'gpay': 'Google Pay',
                'bhim': 'BHIM',
                'ybl': 'PhonePe',
                'okaxis': 'Axis Bank',
                'okhdfcbank': 'HDFC Bank',
                'okicici': 'ICICI Bank',
                'oksbi': 'SBI Bank'
            }
            provider = providers.get(provider_code, provider_code.upper())
    
    return jsonify({
        'valid': is_valid,
        'provider': provider,
        'message': f'Valid {provider} UPI ID' if is_valid else 'Invalid UPI ID format'
    })

@app.route('/welcome')
def welcome():
    name = request.args.get('name', 'Guest')
    
    return f"<h1>Welcome {name}!</h1><p><a href='/dashboard'>Go to Dashboard</a></p>"

if __name__ == '__main__':
    init_excel_files()
    app.run(debug=True, host='0.0.0.0', port=5001)